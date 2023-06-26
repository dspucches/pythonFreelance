# import necessary libraries
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# send a get request to the website
url = "https://www.abuseipdb.com/check/117.211.46.174"
response = requests.get(url)

# parse html content using BeautifulSoup
soup = BeautifulSoup(response.content, "html.parser")

# extract number of reports from page
num_reports = int(soup.find_all(class_="text-muted")[0].text.split()[0])

# create a new workbook and sheet
wb = Workbook()
ws = wb.active

# write data to worksheet
ws['A1'] = 'IP'
ws['B1'] = 'Number of Reports'
ws['A2'] = url
ws['B2'] = num_reports

# save and close workbook
wb.save('abuseipdb_report.xlsx')
