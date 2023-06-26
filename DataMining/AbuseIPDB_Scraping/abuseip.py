# Author: Dominic Spucches
# Date: 6/20/2023
# About: These script was built to input a set of IP addresses then gathers the required data from the output. This includes the number of times that IP has been reported; the domain it's associated with; and the report table. This will output the data into a CSV file.

import time
import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


def uploadform(event=None):
    filename = filedialog.askopenfilename()  # stores imported file into the variable 'filename'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    for cell in sheet['A']:
        value = cell.value

        browser = webdriver.Edge()  # Sets the browser to Microsoft Edge
        browser.get('http://www.abuseipdb.com/')  # opens the edge browser with abuseipdb.com
        assert 'AbuseIPDB' in browser.title

        elem = browser.find_element(By.CLASS_NAME, 'form-control')  # Find the search box by html class name
        elem.send_keys(value + Keys.RETURN)  # inserts the IP address then hits the return key to search

        elem = browser.find_elements(By.ID, 'report-wrapper')  # finds the entire report of the search
        if 'was found in our database!' in elem[0].text:
            elem = browser.find_element(By.CLASS_NAME, 'text-primary')
            workbook = openpyxl.load_workbook('hello.xlsx')
            sheet = workbook.active
            sheet['A1'] = 'Malicious IPs'
            # Set font size to 18 and make text bold
            font = Font(size=18, bold=True)
            sheet['A1'].font = font
            # Find the next available cell in column A
            max_row = sheet.max_row
            next_cell = 'A{}'.format(max_row + 1)
            # Write the value to the next available cell in column A
            sheet[next_cell] = elem.text

            workbook.save('hello.xlsx')
        else:
            print(value + " was not identified as malicious")

        time.sleep(10)
        browser.quit()
        # driver.implicitly_wait(35)


root = tk.Tk()
button = tk.Button(root, text='Import Excel Document', command=uploadform, height=5, width=45)
button.pack()
root.mainloop()
